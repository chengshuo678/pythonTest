# 处理Excel
import openpyxl
wb = openpyxl.load_workbook('test.xlsx')
print(type(wb))
print(wb.get_sheet_names());

sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'].value)
print(sheet.cell(row=2,column=1).value)
for i in range(1,4):
    print(sheet.cell(row=i,column=1).value)
    sheet.cell(row=i,column=2).value=sheet.cell(row=i,column=1).value

for i in range(1,4):
    print(sheet.cell(row=i,column=2).value)

wb.save('test.xlsx')

